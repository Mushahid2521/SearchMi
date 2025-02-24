import openpyxl


def create_search_result_track_output_genetic_algorithm(sa_tracking_dict, file_path):
    wb = openpyxl.Workbook()

    sheet1 = wb.active
    sheet1.title = "Best solution"

    last_generation = max(sa_tracking_dict.keys())
    best_solution = sa_tracking_dict.get(last_generation)
    sheet1.cell(row=1, column=1).value = "Best solution:"
    sheet1.cell(row=2, column=1).value = f"P-value: {best_solution['best_score']}"

    for row_idx, species in enumerate(sorted(best_solution['best_solution']), start=3):
        sheet1.cell(row=row_idx, column=1).value = species

    sheet2 = wb.create_sheet(title="Generations")

    sheet2.cell(row=1, column=1).value = "Generation"
    sheet2.cell(row=1, column=2).value = "p-value"
    row_index = 2

    reverse_index = sorted(sa_tracking_dict.keys(), reverse=True)
    for gen in reverse_index:
        details = sa_tracking_dict.get(gen)

        sheet2.cell(row=row_index, column=1).value = gen
        sheet2.cell(row=row_index, column=2).value = details["best_score"]

        sorted_solution = sorted(details['best_solution'])

        col_index = 3
        for species in sorted_solution:
            sheet2.cell(row=row_index, column=col_index).value = species
            col_index += 1

        row_index += 1

    wb.save(file_path)


def create_search_result_track_output_simulated_annealing(sa_tracking_dict, file_path):
    wb = openpyxl.Workbook()

    sheet1 = wb.active
    sheet1.title = "Best solution"

    sorted_tracking = sorted(sa_tracking_dict.values(), key=lambda x: x['best_score'])
    list_of_lists = [[x['best_score'], x['current_solution']] for x in sorted_tracking]

    best_solution = list_of_lists[0]
    sheet1.cell(row=1, column=1).value = "Best solution:"
    sheet1.cell(row=2, column=1).value = f"P-value: {best_solution[0]}"

    for row_idx, species in enumerate(sorted(best_solution[1]), start=3):
        sheet1.cell(row=row_idx, column=1).value = species

    sheet2 = wb.create_sheet(title="Iterations")

    sheet2.cell(row=1, column=1).value = "Iteration"
    sheet2.cell(row=1, column=2).value = "p-value"
    row_index = 2

    top_k = 20
    for n, details in enumerate(list_of_lists[:top_k]):
        sheet2.cell(row=row_index, column=1).value = f'top {n}'
        sheet2.cell(row=row_index, column=2).value = details[0]

        sorted_solution = sorted(details[1])

        col_index = 3
        for species in sorted_solution:
            sheet2.cell(row=row_index, column=col_index).value = species
            col_index += 1

        row_index += 1

    wb.save(file_path)
